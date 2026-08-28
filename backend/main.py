"""
Onboarding SLA Dashboard - Backend API
Reads the onboarding tracker CSV (uploaded by the user, not a hardcoded local
path) and computes everything the dashboard needs directly from the data
itself - no hardcoded day-count targets. "Normal" for each stage is defined
by the stage's own global average and spread, computed fresh each time.

Rounding policy: every internal computation (averages, std dev, anomaly
cutoffs, deviations, z-scores, focus-area thresholds) is done on raw,
unrounded floats so comparisons stay accurate. Values are only rounded to
whole numbers at the very end, right before they go into a JSON response -
see `round_out()`.
"""
import io
import re
import time
from pathlib import Path

import pandas as pd
from fastapi import FastAPI, UploadFile, File, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse

APP_DIR = Path(__file__).parent
DATA_PATH = APP_DIR / "data" / "onboarding.csv"

app = FastAPI(title="Onboarding SLA Dashboard API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Tracks whatever the user has uploaded this run, so the frontend can show
# "currently loaded: <file>" instead of a hidden local path.
DATASET_INFO = {
    "filename": "onboarding.csv (bundled sample)",
    "uploadedAt": None,
}

# ---------------------------------------------------------------------------
# Config: each "stage" is an already-computed elapsed-day column from the
# source CSV. No fixed targets here - anomalies are judged against each
# stage's own global average + spread, recomputed from the live data.
# ---------------------------------------------------------------------------
STAGE_CONFIG = [
    {"column": "SLA Resource Fulfilment", "label": "Resource Fulfilment"},
    {"column": "SLA - BGV Completion", "label": "BGV Completion"},
    {"column": "SLA Finra Start Date", "label": "FINRA Initiation Gap"},
    {"column": "SLA - Finra Start to End", "label": "FINRA Start to Courier End"},
    {"column": "SLA - Onboarding Docs Submission", "label": "Onboarding Docs Submission"},
    {"column": "SLA - Magnit Documentation", "label": "Magnit Documentation"},
    {"column": "Onboarding Documents Submission - PID", "label": "Docs Submission to PID"},
    {"column": "SLA - MAC Set up", "label": "MAC Setup"},
]

# Purely a presentation grouping of the 8 stages above into the 3 major
# onboarding phases - does not change how any individual stage is
# calculated. Every label here must match a STAGE_CONFIG label exactly.
STAGE_GROUPS = [
    {
        "name": "Resource Requirements to Identification",
        "stages": ["Resource Fulfilment"],
    },
    {
        "name": "Identification to Onboarding",
        "stages": [
            "BGV Completion",
            "FINRA Initiation Gap",
            "FINRA Start to Courier End",
            "Onboarding Docs Submission",
            "Magnit Documentation",
            "Docs Submission to PID",
        ],
    },
    {
        "name": "PID to Billing",
        "stages": ["MAC Setup"],
    },
]

GROUP_BY_COLUMN = "Project Details"
TOTAL_DURATION_COLUMN = "Complete Onboarding"
NAME_COLUMN = "Name"
COMPLETION_DATE_COLUMN = "Onboarding Completion Date"

FILTER_COLUMNS = ["Project Details", "Location", "Status", "CGI/External", "BGV Status"]

# A stage day-count can never be negative, and a gap over ~1 year almost
# certainly means the CSV cell had something other than a plain number in it
# (a stray date, a formula artifact, a typo). Anything outside this range is
# treated as bad data: excluded from averages and reported, not silently
# averaged in.
MIN_PLAUSIBLE_DAYS = 0
MAX_PLAUSIBLE_DAYS = 365

# How far above the global average counts as an anomaly, in standard
# deviations. 1.0 = flag anything meaningfully worse than typical.
ANOMALY_Z = 1.0

# A candidate is a "focus area" for a stage once they've used up this
# fraction of the stage's own global average - i.e. approaching, but not
# yet past, the anomaly cutoff. Only candidates who haven't completed
# onboarding yet are eligible (there's nothing to "focus on" for someone
# already done), and anything past the anomaly cutoff is excluded here so a
# candidate is never counted as both a focus area and an anomaly for the
# same stage.
FOCUS_AREA_THRESHOLD = 0.75


def round_out(value):
    """Round a raw numeric value to a whole number for JSON output. Returns
    None for missing/NaN values so the frontend's existing null-handling
    ('—' placeholders etc.) keeps working unchanged."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return int(round(float(value)))


def normalize(name: str) -> str:
    """Collapse repeated whitespace so header quirks (double spaces, etc.)
    in the source CSV don't break column lookups."""
    return re.sub(r"\s+", " ", str(name).strip())


def clean_numeric_column(series: pd.Series) -> tuple[pd.Series, int]:
    """Coerce to numeric and drop implausible values (negative or absurdly
    large), which are near-certainly parsing artifacts rather than real
    elapsed-day counts. Returns the cleaned series and a count of rows
    dropped for that reason (separate from rows that were already blank)."""
    numeric = pd.to_numeric(series, errors="coerce")
    was_present = numeric.notna()
    implausible = was_present & ((numeric < MIN_PLAUSIBLE_DAYS) | (numeric > MAX_PLAUSIBLE_DAYS))
    dropped = int(implausible.sum())
    numeric = numeric.where(~implausible, other=pd.NA)
    return numeric, dropped


def load_dataframe(path: Path = DATA_PATH) -> tuple[pd.DataFrame, dict]:
    if not path.exists():
        raise HTTPException(
            status_code=404,
            detail="No dataset loaded yet. Upload a CSV via the dashboard's upload button.",
        )

    df = pd.read_csv(path)
    df.columns = [normalize(c) for c in df.columns]

    data_quality = {}
    for stage in STAGE_CONFIG:
        col = stage["column"]
        if col in df.columns:
            cleaned, dropped = clean_numeric_column(df[col])
            df[col] = cleaned
            if dropped:
                data_quality[stage["label"]] = dropped

    if TOTAL_DURATION_COLUMN in df.columns:
        df[TOTAL_DURATION_COLUMN] = pd.to_numeric(df[TOTAL_DURATION_COLUMN], errors="coerce")

    for col in FILTER_COLUMNS:
        if col in df.columns:
            df[col] = df[col].fillna("Unspecified").astype(str).str.strip()
            df[col] = df[col].replace({"": "Unspecified", "nan": "Unspecified"})

    return df, data_quality


def get_active_filters(request: Request) -> dict:
    """Reads any of FILTER_COLUMNS present as query params, e.g.
    ?Project Details=Mobile Team&Status=Active. Empty/missing values are
    ignored so 'no filter' and 'filter cleared' behave the same."""
    filters = {}
    for col in FILTER_COLUMNS:
        val = request.query_params.get(col)
        if val:
            filters[col] = val
    return filters


def apply_filters(df: pd.DataFrame, filters: dict) -> pd.DataFrame:
    for col, val in filters.items():
        if col in df.columns:
            df = df[df[col] == val]
    return df


def is_onboarding_complete(row: pd.Series) -> bool:
    """A candidate counts as 'done' if they have a completion date, or -
    falling back for datasets without that column - a non-null total
    onboarding duration."""
    if COMPLETION_DATE_COLUMN in row.index:
        val = row.get(COMPLETION_DATE_COLUMN)
        if pd.notna(val) and str(val).strip() != "":
            return True
        return False
    if TOTAL_DURATION_COLUMN in row.index:
        return pd.notna(row.get(TOTAL_DURATION_COLUMN))
    return False


def stage_stats(df: pd.DataFrame) -> dict:
    """Global average + std dev per stage, computed fresh from the cleaned
    data, kept as raw (unrounded) floats - this is the whole basis for what
    counts as an anomaly or a focus area, and rounding it early would let
    borderline cases fall on the wrong side of a cutoff. When filters are
    active, `df` is already narrowed to the selected subset, so stats
    reflect that subset rather than the whole dataset."""
    stats = {}
    for stage in STAGE_CONFIG:
        col = stage["column"]
        if col not in df.columns:
            continue
        valid = df[col].dropna()
        if valid.empty:
            stats[stage["label"]] = {"average": None, "std": None, "anomalyCutoff": None}
            continue
        avg = float(valid.mean())
        std = float(valid.std()) if len(valid) > 1 else 0.0
        stats[stage["label"]] = {
            "average": avg,
            "std": std,
            "anomalyCutoff": avg + ANOMALY_Z * std,
        }
    return stats


def build_candidate_records(df: pd.DataFrame, stats: dict) -> list[dict]:
    records = []
    for _, row in df.iterrows():
        record = {}
        for col in df.columns:
            val = row[col]
            if pd.isna(val):
                record[col] = None
            elif isinstance(val, float) and val.is_integer():
                record[col] = int(val)
            else:
                record[col] = val

        anomalies = {}
        stage_values = {}
        for stage in STAGE_CONFIG:
            col = stage["column"]
            val = row.get(col)
            label = stage["label"]
            cutoff = stats.get(label, {}).get("anomalyCutoff")
            if pd.isna(val):
                anomalies[label] = None
                stage_values[label] = None
            else:
                stage_values[label] = round_out(val)
                anomalies[label] = bool(cutoff is not None and val > cutoff)
        record["_anomalies"] = anomalies
        record["_stageValues"] = stage_values
        records.append(record)
    return records


def build_anomaly_watchlist(df: pd.DataFrame, stats: dict) -> list[dict]:
    """Flat, ranked list of every candidate/stage anomaly - this is the
    'what exactly is wrong' view, worst deviation first."""
    watchlist = []
    name_col = NAME_COLUMN if NAME_COLUMN in df.columns else None
    group_col = GROUP_BY_COLUMN if GROUP_BY_COLUMN in df.columns else None

    for _, row in df.iterrows():
        for stage in STAGE_CONFIG:
            col = stage["column"]
            label = stage["label"]
            info = stats.get(label, {})
            avg = info.get("average")
            std = info.get("std")
            cutoff = info.get("anomalyCutoff")
            val = row.get(col)
            if col not in df.columns or pd.isna(val) or cutoff is None or val <= cutoff:
                continue
            deviation = float(val) - avg
            z_score = deviation / std if std else None
            watchlist.append({
                "candidate": str(row[name_col]) if name_col else "Unknown",
                "team": str(row[group_col]) if group_col else "Unspecified",
                "stage": label,
                "value": round_out(val),
                "average": round_out(avg),
                "deviation": round_out(deviation),
                "zScore": round_out(z_score),
                "_sortDeviation": deviation,
            })

    watchlist.sort(key=lambda x: x["_sortDeviation"], reverse=True)
    for entry in watchlist:
        del entry["_sortDeviation"]
    return watchlist


def build_focus_areas(df: pd.DataFrame, stats: dict) -> list[dict]:
    """Candidates who are trending toward - but haven't yet crossed - a
    stage's anomaly cutoff, so this can act as an early-warning list rather
    than an after-the-fact one. Only candidates still mid-onboarding are
    eligible (nothing to focus on once someone's finished), and anything at
    or past the anomaly cutoff is excluded so a candidate never shows up as
    both a focus area and an anomaly for the same stage."""
    focus_areas = []
    name_col = NAME_COLUMN if NAME_COLUMN in df.columns else None
    group_col = GROUP_BY_COLUMN if GROUP_BY_COLUMN in df.columns else None

    for _, row in df.iterrows():
        if is_onboarding_complete(row):
            continue
        for stage in STAGE_CONFIG:
            col = stage["column"]
            label = stage["label"]
            info = stats.get(label, {})
            avg = info.get("average")
            cutoff = info.get("anomalyCutoff")
            val = row.get(col)
            if col not in df.columns or pd.isna(val) or avg is None or cutoff is None or avg <= 0:
                continue
            threshold = FOCUS_AREA_THRESHOLD * avg
            if val < threshold or val > cutoff:
                continue
            percent_of_average = (float(val) / avg) * 100
            focus_areas.append({
                "candidate": str(row[name_col]) if name_col else "Unknown",
                "team": str(row[group_col]) if group_col else "Unspecified",
                "stage": label,
                "value": round_out(val),
                "average": round_out(avg),
                "percentOfAverage": round_out(percent_of_average),
                "_sortPercent": percent_of_average,
            })

    focus_areas.sort(key=lambda x: x["_sortPercent"], reverse=True)
    for entry in focus_areas:
        del entry["_sortPercent"]
    return focus_areas


def build_stage_groups(stage_summary: list[dict]) -> list[dict]:
    """Rolls the flat per-stage summary up into the 3 major onboarding
    phases defined in STAGE_GROUPS. Purely a presentation aggregation - the
    underlying per-stage averages/anomaly counts (computed elsewhere) are
    untouched, this just sums them per phase:
      - a phase's 'average' is the sum of its sub-stages' averages (each
        sub-stage average already represents days spent in that phase, so
        summing gives total days for the phase as a whole)
      - a phase's 'anomalyCount' is the sum of its sub-stages' anomaly
        counts
    """
    by_label = {s["label"]: s for s in stage_summary}
    groups = []
    for group in STAGE_GROUPS:
        sub_stages = [by_label[label] for label in group["stages"] if label in by_label]
        averages = [s["average"] for s in sub_stages if s["average"] is not None]
        groups.append({
            "name": group["name"],
            "average": int(round(sum(averages))) if averages else None,
            "anomalyCount": sum(s["anomalyCount"] for s in sub_stages),
            "subStages": sub_stages,
        })
    return groups


@app.get("/api/candidates")
def get_candidates(request: Request):
    full_df, _ = load_dataframe()
    df = apply_filters(full_df, get_active_filters(request))
    stats = stage_stats(df)
    return {"columns": list(df.columns), "rows": build_candidate_records(df, stats)}


@app.get("/api/summary")
def get_summary(request: Request):
    """Global average per stage, team (Project Details) averages against
    that global average, anomaly counts, a ranked anomaly watchlist, and a
    ranked focus-areas list - all derived from the (optionally filtered)
    data, no fixed targets.

    Filter dropdown option lists (`filters` in the response) are always
    built from the FULL, unfiltered dataset so previously-available options
    don't disappear as other filters get applied.
    """
    full_df, data_quality = load_dataframe()
    active_filters = get_active_filters(request)
    df = apply_filters(full_df, active_filters)
    stats = stage_stats(df)

    groups = df[GROUP_BY_COLUMN].unique().tolist() if GROUP_BY_COLUMN in df.columns else ["All"]

    by_group = []
    for group in groups:
        sub = df[df[GROUP_BY_COLUMN] == group] if GROUP_BY_COLUMN in df.columns else df
        entry = {"group": group, "candidateCount": int(len(sub))}
        for stage in STAGE_CONFIG:
            col = stage["column"]
            label = stage["label"]
            if col not in df.columns:
                continue
            raw_team_avg = sub[col].mean()
            team_avg = None if pd.isna(raw_team_avg) else float(raw_team_avg)
            entry[label] = round_out(team_avg)
            global_avg = stats.get(label, {}).get("average")
            entry[f"{label}__vsAvg"] = (
                round_out(team_avg - global_avg) if team_avg is not None and global_avg is not None else None
            )
        by_group.append(entry)

    stage_summary = []
    total_anomalies = 0
    for stage in STAGE_CONFIG:
        col = stage["column"]
        label = stage["label"]
        if col not in df.columns:
            continue
        valid = df[col].dropna()
        cutoff = stats.get(label, {}).get("anomalyCutoff")
        anomaly_count = int((valid > cutoff).sum()) if cutoff is not None else 0
        total_anomalies += anomaly_count
        stage_summary.append({
            "label": label,
            "column": col,
            "average": round_out(stats.get(label, {}).get("average")),
            "std": round_out(stats.get(label, {}).get("std")),
            "anomalyCutoff": round_out(cutoff),
            "anomalyCount": anomaly_count,
            "trackedCount": int(len(valid)),
            "excludedBadData": data_quality.get(label, 0),
        })

    total_duration_valid = (
        df[TOTAL_DURATION_COLUMN].dropna()
        if TOTAL_DURATION_COLUMN in df.columns
        else pd.Series(dtype=float)
    )

    kpis = {
        "totalCandidates": int(len(df)),
        "totalAnomalies": total_anomalies,
        "projectGroups": int(df[GROUP_BY_COLUMN].nunique()) if GROUP_BY_COLUMN in df.columns else 0,
        "avgTotalOnboardingDays": round_out(total_duration_valid.mean()) if not total_duration_valid.empty else None,
        "minTotalOnboardingDays": round_out(total_duration_valid.min()) if not total_duration_valid.empty else None,
        "maxTotalOnboardingDays": round_out(total_duration_valid.max()) if not total_duration_valid.empty else None,
        "dataQualityIssues": sum(data_quality.values()),
    }

    filters = {}
    for col in FILTER_COLUMNS:
        if col in full_df.columns:
            filters[col] = sorted(full_df[col].unique().tolist())

    return {
        "kpis": kpis,
        "stages": stage_summary,
        "stageGroups": build_stage_groups(stage_summary),
        "byGroup": by_group,
        "groupColumn": GROUP_BY_COLUMN,
        "filters": filters,
        "activeFilters": active_filters,
        "dataQuality": data_quality,
        "anomalyWatchlist": build_anomaly_watchlist(df, stats),
        "focusAreas": build_focus_areas(df, stats),
        "dataset": DATASET_INFO,
    }


@app.get("/api/export")
def export_excel():
    df, _ = load_dataframe()
    stats = stage_stats(df)
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Onboarding Data")
        worksheet = writer.sheets["Onboarding Data"]

        from openpyxl.styles import PatternFill

        anomaly_fill = PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid")
        header_map = {cell.value: cell.column for cell in worksheet[1]}

        for stage in STAGE_CONFIG:
            col_name = stage["column"]
            cutoff = stats.get(stage["label"], {}).get("anomalyCutoff")
            if col_name not in header_map or cutoff is None:
                continue
            col_idx = header_map[col_name]
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)) and cell.value > cutoff:
                    cell.fill = anomaly_fill

    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Onboarding_SLA_Data.xlsx"},
    )


@app.post("/api/upload")
async def upload_csv(file: UploadFile = File(...)):
    """Replace the working dataset with a user-uploaded CSV. This is the
    only way data gets into the app - no local file path required."""
    if not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Please upload a .csv file.")

    contents = await file.read()
    DATA_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(DATA_PATH, "wb") as f:
        f.write(contents)

    # Sanity check it actually parses before reporting success
    try:
        df = pd.read_csv(io.BytesIO(contents))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Couldn't parse that CSV: {exc}")

    DATASET_INFO["filename"] = file.filename
    DATASET_INFO["uploadedAt"] = time.time()

    return {"status": "ok", "filename": file.filename, "rowCount": int(len(df))}


@app.get("/api/dataset-info")
def dataset_info():
    return DATASET_INFO


@app.get("/api/health")
def health():
    return {"status": "ok"}